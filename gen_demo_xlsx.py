import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Demo_Bulk_Upload'

header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

headers = ['Question_id', 'Technology_Stack', 'Topic', 'Difficulty', 'Question_Stem', 'Option_A', 'Option_B', 'Option_C', 'Option_D', 'Correct_Answer']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

rows = [
    # --- SEMANTIC DUPLICATES (will be BLOCKED) ---
    # Row 2 - STRONGER dup of MCQ#2001 "What is Dependency Injection in Spring and how does it promote loose coupling?"
    [1, 'Spring Core', 'Dependency Injection', 'MEDIUM',
     'How does Spring implement Dependency Injection to achieve loose coupling between components?',
     'Through tight coupling', 'Container injects dependencies so classes dont create their own reducing coupling',
     'Through manual instantiation', 'Through static methods only', 'B'],

    # Row 3 - DUP of MCQ#2002 "How does the Spring IoC container manage object creation and inject dependencies?"
    [2, 'Spring Core', 'Dependency Injection', 'EASY',
     'Explain how IoC container manages beans and wires dependencies automatically',
     'It compiles code', 'It creates objects and injects their dependencies automatically',
     'It deploys apps', 'It manages HTTP sessions', 'B'],

    # Row 4 - DUP of MCQ#2005 "What is Aspect-Oriented Programming (AOP) in Spring and what problems does it solve?"
    [3, 'Spring Core', 'AOP', 'MEDIUM',
     'How does Spring use Aspect Oriented Programming for cross-cutting concerns like logging?',
     'Through manual coding', 'Through aspects that intercept at pointcuts for logging and transactions',
     'Through database views', 'Through config files', 'B'],

    # Row 5 - DUP of MCQ#2010 "What is runtime polymorphism in Java and how is it achieved through method overriding?"
    [4, 'Core Java', 'OOP Concepts', 'MEDIUM',
     'How does method overriding enable runtime polymorphism in Java applications?',
     'Through static binding', 'Subclass provides specific implementation selected dynamically at runtime',
     'Through constructors', 'Through generics', 'B'],

    # Row 6 - DUP of MCQ#2012 "How do you create a thread in Java and what is the difference between Thread and Runnable?"
    [5, 'Core Java', 'Multithreading', 'EASY',
     'What are the ways to create threads in Java using Thread class vs Runnable interface?',
     'No difference', 'Thread extends class but limits inheritance while Runnable allows more flexibility',
     'Both are deprecated', 'Only Thread class works', 'B'],

    # Row 7 - DUP of MCQ#2006 "How does Spring Boot auto-configuration work and which annotation enables it?"
    [6, 'Spring Boot', 'Auto Configuration', 'MEDIUM',
     'How does Spring Boot auto-configuration work internally to configure beans?',
     'Downloads config from cloud', 'Reads META-INF/spring.factories and applies conditional configuration',
     'Requires manual XML', 'Uses bytecode manipulation', 'B'],

    # --- NEW UNIQUE QUESTIONS (will PASS) ---
    # Row 8 - Truly unique: Java Streams (not in DB at all)
    [7, 'Core Java', 'Streams API', 'MEDIUM',
     'How does the Java Stream API support functional-style operations on collections?',
     'By modifying the original collection', 'Through lazy evaluation pipelines with map filter and reduce operations',
     'By converting to arrays first', 'Through reflection only', 'B'],

    # Row 9 - Truly unique: Spring Security (not in DB)
    [8, 'Spring Boot', 'Security', 'HARD',
     'How does Spring Security implement authentication and authorization using filter chains?',
     'Through database triggers', 'Using servlet filters that intercept requests and validate credentials before reaching controllers',
     'Through URL rewriting', 'By modifying HTTP headers only', 'B'],

    # Row 10 - Truly unique: Garbage Collection (not in DB)
    [9, 'Core Java', 'Memory Management', 'HARD',
     'How does the JVM garbage collector reclaim unused memory and what are the different GC algorithms?',
     'Manual memory free calls', 'Through mark-and-sweep and generational collection with Young and Old generations',
     'Through reference counting only', 'By restarting the JVM', 'B'],

    # Row 11 - Truly unique: Design Patterns (not in DB)
    [10, 'Core Java', 'Design Patterns', 'MEDIUM',
     'What is the Singleton design pattern and how do you implement it thread-safely in Java?',
     'Create multiple instances', 'Use private constructor with synchronized getInstance or enum-based implementation',
     'Use public constructors', 'Singletons cannot be thread-safe', 'B'],
]

dup_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
new_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

for i, row_data in enumerate(rows, 2):
    fill = dup_fill if i <= 7 else new_fill  # rows 2-7 yellow, rows 8-11 green
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=i, column=col, value=val)
        cell.border = thin_border
        cell.fill = fill
        if col == 5:
            cell.alignment = Alignment(wrap_text=True)

col_widths = [12, 20, 22, 12, 80, 50, 55, 45, 45, 15]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# README sheet
ws2 = wb.create_sheet('README')
ws2['A1'] = 'DEMO: AI-Powered Semantic Duplicate Detection in Bulk Upload'
ws2['A1'].font = Font(bold=True, size=14)
ws2['A3'] = 'Yellow rows (1-6): SEMANTIC DUPLICATES - same concept, different wording -> BLOCKED by AI'
ws2['A4'] = 'Green rows (7-10): GENUINELY NEW QUESTIONS - unique topics -> IMPORTED successfully'
ws2['A6'] = 'Expected Demo Results:'
ws2['A7'] = '  Row 2: "How does Spring implement DI for loose coupling..." -> matches MCQ#2001 (DI question)'
ws2['A8'] = '  Row 3: "Explain how IoC container manages beans..." -> matches MCQ#2002 (IoC question)'
ws2['A9'] = '  Row 4: "How does Spring use AOP for cross-cutting..." -> matches MCQ#2005 (AOP question)'
ws2['A10'] = '  Row 5: "How does method overriding enable polymorphism..." -> matches MCQ#2010'
ws2['A11'] = '  Row 6: "Ways to create threads - Thread vs Runnable..." -> matches MCQ#2012'
ws2['A12'] = '  Row 7: "How does Spring Boot auto-config work..." -> matches MCQ#2006'
ws2['A13'] = '  Row 8-11: Java Streams, Spring Security, GC, Singleton -> NEW topics, will PASS'
ws2['A15'] = 'Demo Steps:'
ws2['A16'] = '  1. Login as SME (birendra.kumar.singh / Sme@1234)'
ws2['A17'] = '  2. Go to Bulk Upload page'
ws2['A18'] = '  3. Upload this file -> watch AI analyze each question'
ws2['A19'] = '  4. Show results: 6 BLOCKED (semantic) + 4 IMPORTED (new)'
ws2['A20'] = '  5. Click "View" on blocked row -> side-by-side comparison with existing MCQ'
ws2['A21'] = '  6. Click "Add Anyway" -> override AI decision (human in the loop)'
ws2['A22'] = '  7. Click "Edit & Submit" -> modify question to make it unique'
ws2.column_dimensions['A'].width = 90

wb.save('Demo_BulkUpload_SemanticDuplicates.xlsx')
print('SUCCESS: Demo_BulkUpload_SemanticDuplicates.xlsx created')
print('  6 yellow rows = semantic duplicates (will be BLOCKED by AI)')
print('  4 green rows = new unique questions (will be IMPORTED)')
